# Copyright (c) Microsoft Corporation.
# Licensed under the MIT License.

"""Core SheetBrain agent with three-stage architecture."""

import os
import time
from typing import Dict, Any, Optional

from PIL import Image
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config.settings import Config
from modules.understanding import UnderstandingModule
from modules.execution import ExecutionModule
from modules.validation import ValidationModule
from utils.excel_toolkit import ExcelToolkit, calculate_token_cost_line
from utils.logger import setup_logger

logger = setup_logger(__name__)


class SheetBrain:
    """
    Excel analysis agent with three-stage architecture: Understand-Execute-Validate.
    Supports iterative improvement through validation feedback.
    """

    def __init__(self, excel_path: str, config: Optional[Config] = None,
                 total_token_budget: int = 10000,
                 load_excel: bool = True, excel_context_understanding: Optional[str] = None,
                 excel_context_execution: Optional[str] = None):
        """
        Initialize the SheetBrain agent.

        Args:
            excel_path: Path to the Excel file
            config: Configuration object (if None, uses default config)
            total_token_budget: Token budget for initial context generation
            load_excel: Whether to load Excel file and openpyxl libraries
            excel_context_understanding: Pre-generated understanding context (optional)
            excel_context_execution: Pre-generated execution context (optional)
        """
        self.excel_path = excel_path
        self.config = config or Config()
        self.total_token_budget = total_token_budget
        self.load_excel = load_excel

        # Initialize OpenAI client
        self.client = OpenAI(
            api_key=self.config.api_key,
            base_url=self.config.base_url
        )

        # Initialize code execution environment
        self.code_globals = {
            'math': __import__('math'),
            'json': __import__('json'),
            're': __import__('re'),
            'os': __import__('os'),
            'sys': __import__('sys'),
            'excel_path': excel_path,
        }
        self.code_locals = {}

        # Setup Excel environment
        if self.load_excel:
            self._setup_excel_libraries()
            self.workbook = self.code_globals['workbook']

            # Generate Excel context
            if excel_context_understanding is None:
                self.excel_context_understanding = self._generate_sheets_markdown_summary(total_token_budget * 2)
            else:
                self.excel_context_understanding = excel_context_understanding

            if excel_context_execution is None:
                self.excel_context_execution = self._generate_sheets_markdown_summary(total_token_budget)
            else:
                self.excel_context_execution = excel_context_execution
        else:
            self.workbook = None
            self.excel_context_understanding = excel_context_understanding or "Excel file not loaded. Working with provided context only."
            self.excel_context_execution = excel_context_execution or "Excel file not loaded. Working with provided context only."

        # Initialize the three modules
        self.understanding_module = UnderstandingModule(
            self.client, self.config.deployment, self.excel_context_understanding, self.workbook
        )
        self.execution_module = ExecutionModule(
            self.client, self.config.deployment, self.code_globals, self.code_locals, self.excel_context_execution
        )
        self.validation_module = ValidationModule(
            self.client, self.config.deployment, self.excel_context_understanding
        )

    def run(self, user_question: str, table_image: Optional[Image.Image] = None,
            max_turns: Optional[int] = None, enable_validation: Optional[bool] = None,
            enable_understanding: Optional[bool] = None) -> Dict[str, Any]:
        """
        Run the complete analysis with iterative improvement through validation.

        Args:
            user_question: The user's query about the Excel file
            table_image: Optional screenshot of relevant Excel sheet area
            max_turns: Maximum number of Execute-Validate iterations
            enable_validation: Whether to run the validation stage
            enable_understanding: Whether to run the understanding stage

        Returns:
            Dictionary containing complete analysis results
        """
        # Use config defaults if not specified
        max_turns = max_turns or self.config.max_turns
        enable_validation = enable_validation if enable_validation is not None else self.config.enable_validation
        enable_understanding = enable_understanding if enable_understanding is not None else self.config.enable_understanding

        logger.info("Starting iterative three-stage analysis")
        print("🚀 [SheetBrain] Starting iterative three-stage analysis...")
        print("="*80)

        overall_start_time = time.time()
        all_execution_results = []
        all_validation_results = []

        try:
            # ===== STAGE 1: UNDERSTANDING (Optional) =====
            if enable_understanding:
                logger.info("Running understanding module")
                print("📖 [STAGE 1] UNDERSTANDING MODULE")
                print("-" * 40)
                understanding_start_time = time.time()

                understanding_output = self.understanding_module.analyze(user_question, table_image)
                understanding_duration = time.time() - understanding_start_time

                print(f"✅ [STAGE 1] Understanding completed in {understanding_duration:.2f}s")
                print(f"📝 [STAGE 1] Analysis preview: {understanding_output}...")
            else:
                logger.info("Understanding module disabled")
                print("⏭️ [STAGE 1] UNDERSTANDING MODULE SKIPPED")
                print("-" * 40)
                understanding_output = f"Understanding module disabled. Direct analysis of user question: {user_question}"
                print(f"📝 [STAGE 1] Using direct question: {user_question}")

            # ===== ITERATIVE EXECUTE-VALIDATE LOOP =====
            for iteration in range(max_turns):
                logger.info(f"Starting iteration {iteration + 1}/{max_turns}")
                print(f"\n🔄 [ITERATION {iteration + 1}/{max_turns}] EXECUTE-VALIDATE CYCLE")
                print("="*60)

                # ===== STAGE 2: EXECUTION =====
                print(f"💻 [ITERATION {iteration + 1}] EXECUTION MODULE")
                print("-" * 40)
                execution_start_time = time.time()

                # Add improvement feedback from previous validation if available
                if iteration > 0 and all_validation_results:
                    last_validation = all_validation_results[-1]
                    if last_validation.get('improvement_feedback'):
                        enhanced_understanding = f"""{understanding_output}

**IMPROVEMENT FEEDBACK FROM PREVIOUS ITERATION:**
{last_validation['improvement_feedback']}

**ISSUES TO ADDRESS:**
{'; '.join(last_validation.get('issues_found', []))}

Please address these specific points in your new analysis approach."""
                    else:
                        enhanced_understanding = understanding_output
                else:
                    enhanced_understanding = understanding_output

                execution_result = self.execution_module.run(enhanced_understanding, user_question)
                execution_duration = time.time() - execution_start_time
                all_execution_results.append(execution_result)

                status_emoji = "✅" if execution_result["success"] else "❌"
                print(f"{status_emoji} [ITERATION {iteration + 1}] Execution completed in {execution_duration:.2f}s")
                print(f"🔄 [ITERATION {iteration + 1}] Total turns: {execution_result['total_turns']}")
                print(f"📊 [ITERATION {iteration + 1}] Code executions: {execution_result.get('execution_summary', {}).get('total_code_executions', 0)}")

                # ===== STAGE 3: VALIDATION (if enabled) =====
                if enable_validation:
                    logger.info(f"Running validation module for iteration {iteration + 1}")
                    print(f"\n🔍 [ITERATION {iteration + 1}] VALIDATION MODULE")
                    print("-" * 40)
                    validation_start_time = time.time()

                    validation_result = self.validation_module.reflect(execution_result, user_question, understanding_output)
                    validation_duration = time.time() - validation_start_time
                    all_validation_results.append(validation_result)

                    validation_emoji = "✅" if validation_result["validation_passed"] else "⚠️"
                    print(f"{validation_emoji} [ITERATION {iteration + 1}] Validation completed in {validation_duration:.2f}s")
                    print(f"🎯 [ITERATION {iteration + 1}] Confidence: {validation_result['confidence_score']:.2f}")
                    print(f"📋 [ITERATION {iteration + 1}] Validation: {'PASSED' if validation_result['validation_passed'] else 'FAILED'}")

                    # Check if we should stop iterating
                    if validation_result['validation_passed']:
                        logger.info(f"Validation passed on iteration {iteration + 1}")
                        print(f"🎉 [SUCCESS] Validation passed on iteration {iteration + 1}!")
                        final_answer = validation_result.get('verified_answer', execution_result['answer'])
                        overall_success = True
                        confidence_score = validation_result['confidence_score']
                        validation_passed = True
                        break
                    elif not validation_result.get('requires_reexecution', True):
                        logger.warning("Validation indicates no further improvement possible")
                        print(f"🛑 [STOPPING] Validation indicates no further improvement possible")
                        final_answer = execution_result['answer']
                        overall_success = False
                        confidence_score = validation_result['confidence_score']
                        validation_passed = False
                        break
                    else:
                        logger.info(f"Issues found, preparing for iteration {iteration + 2}")
                        print(f"🔄 [CONTINUE] Issues found, preparing for iteration {iteration + 2}")
                        if iteration == max_turns - 1:
                            logger.warning("Reached maximum iterations without validation")
                            print(f"⚠️ [MAX ITERATIONS] Reached maximum iterations without validation")
                            final_answer = execution_result['answer']
                            overall_success = False
                            confidence_score = validation_result['confidence_score']
                            validation_passed = False
                else:
                    # No validation, use execution results directly
                    logger.info("Validation disabled, using execution results directly")
                    final_answer = execution_result['answer']
                    overall_success = execution_result['success']
                    confidence_score = 0.8 if execution_result['success'] else 0.3
                    validation_passed = execution_result['success']
                    break

            # Generate final report
            if all_validation_results:
                final_validation = all_validation_results[-1]
                issues_found = final_validation.get('issues_found', [])
                improvement_feedback = final_validation.get('improvement_feedback', '')
            else:
                issues_found = []
                improvement_feedback = ''

            # Collect all conversation histories from execution results
            all_conversation_histories = []
            for exec_result in all_execution_results:
                conv_history = exec_result.get('conversation_history', [])
                if conv_history:
                    all_conversation_histories.append({
                        'iteration': all_execution_results.index(exec_result) + 1,
                        'conversation_history': conv_history
                    })

            # ===== FINAL SUMMARY =====
            total_duration = time.time() - overall_start_time
            total_iterations = len(all_execution_results)

            logger.info(f"Analysis completed. Success: {overall_success}, Iterations: {total_iterations}")
            print("\n" + "="*80)
            print("🎯 [FINAL SUMMARY]")
            print("="*80)
            print(f"Overall Success: {'✅ YES' if overall_success else '❌ NO'}")
            print(f"Total Iterations: {total_iterations}")
            print(f"Final Answer: {final_answer}")
            print(f"Confidence Score: {confidence_score:.2f}/1.0")
            print(f"Validation Passed: {'✅ YES' if validation_passed else '❌ NO'}")
            print(f"Total Duration: {total_duration:.2f}s")
            print("="*80)

            return {
                "success": overall_success,
                "answer": final_answer,
                "confidence_score": confidence_score,
                "validation_passed": validation_passed,
                "total_iterations": total_iterations,
                "all_execution_results": all_execution_results,
                "all_validation_results": all_validation_results,
                "conversation_history": all_conversation_histories,
                "issues_found": issues_found,
                "improvement_feedback": improvement_feedback,
                "total_duration": total_duration,
                "user_question": user_question,
                "understanding_output": understanding_output
            }

        except Exception as e:
            error_duration = time.time() - overall_start_time
            logger.error(f"Critical error: {str(e)}")
            print(f"❌ [SheetBrain] Critical error: {str(e)}")
            print(f"⏱️ [SheetBrain] Failed after {error_duration:.2f}s")

            # Collect conversation histories even in error case
            all_conversation_histories = []
            for exec_result in all_execution_results:
                conv_history = exec_result.get('conversation_history', [])
                if conv_history:
                    all_conversation_histories.append({
                        'iteration': all_execution_results.index(exec_result) + 1,
                        'conversation_history': conv_history
                    })

            return {
                "success": False,
                "answer": f"Analysis failed due to error: {str(e)}",
                "confidence_score": 0.0,
                "validation_passed": False,
                "total_iterations": len(all_execution_results),
                "all_execution_results": all_execution_results,
                "all_validation_results": all_validation_results,
                "conversation_history": all_conversation_histories,
                "issues_found": [f"Critical error: {str(e)}"],
                "improvement_feedback": "Review the error and try again",
                "total_duration": error_duration,
                "user_question": user_question
            }

    def _generate_sheets_markdown_summary(self, total_token_budget: int = 50000) -> str:
        """Generate a markdown summary of all sheets in the workbook."""
        try:
            workbook = self.workbook
            overview_parts = []

            overview_parts.append(f"📊 **Excel File Overview: {os.path.basename(self.excel_path)}**\n")
            overview_parts.append(f"**Total Sheets:** {len(workbook.sheetnames)}\n")

            # Token budget management
            available_tokens = total_token_budget

            # Distribute tokens among sheets
            tokens_per_sheet = available_tokens // len(workbook.sheetnames) if workbook.sheetnames else 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_parts = []

                sheet_parts.append(f"\n**📄 Sheet: '{sheet_name}'**")
                sheet_parts.append(f"- Dimensions: {sheet.max_row} rows × {sheet.max_column} columns")

                if tokens_per_sheet > 0:
                    # Get comprehensive preview with token limit
                    preview_result = self._get_sheet_preview_with_token_limit(
                        sheet,
                        tokens_per_sheet,
                        max_rows=min(sheet.max_row, 10000),  # Cap at 10000 rows for performance
                        max_cols=min(sheet.max_column, 1000)   # Cap at 1000 columns
                    )

                    sheet_parts.append(f"- Data Preview ({preview_result['rows_shown']} of {sheet.max_row} rows, "
                                    f"{preview_result['cols_shown']} of {sheet.max_column} columns):")

                    if preview_result['is_truncated']:
                        sheet_parts.append("  ⚠️ Preview truncated to fit token budget")

                    # Add data preview in markdown table format with A1 notation
                    sheet_parts.append("  Data:")
                    markdown_rows = []
                    for row_data in preview_result['formatted_data']:
                        markdown_rows.append(f"| {' | '.join(row_data)} |")

                    # Join all rows with newline and backslash-n for compact representation
                    if markdown_rows:
                        sheet_parts.append("  " + "\\n".join(markdown_rows))

                    # Add data summary if we couldn't show all rows
                    if preview_result['rows_shown'] < sheet.max_row:
                        sheet_parts.append(f"\n  📊 Sheet Summary:")
                        sheet_parts.append(f"  - Total rows: {sheet.max_row}")
                        sheet_parts.append(f"  - Total columns: {sheet.max_column}")
                        sheet_parts.append(f"  - Rows shown in preview: {preview_result['rows_shown']}")

                overview_parts.extend(sheet_parts)

            final_overview = "\n".join(overview_parts)
            return final_overview

        except Exception as e:
            logger.error(f"Error generating Excel overview: {str(e)}")
            return f"❌ Error generating Excel overview: {str(e)}"

    def _get_sheet_preview_with_token_limit(self, sheet, token_budget: int,
                                          max_rows: int = 10000, max_cols: int = 1000) -> Dict[str, Any]:
        """Get a preview of sheet data that fits within a token budget."""
        preview_data = []
        formatted_data = []
        tokens_used = 0
        rows_shown = 0

        start_row = 1

        # Calculate effective limits
        max_data_rows = min(max_rows, sheet.max_row)
        max_data_cols = min(max_cols, sheet.max_column)

        # Iterate through rows and accumulate data within token budget
        for row_idx in range(start_row, max_data_rows + 1):
            row_cells = []
            formatted_row_cells = []

            # Get actual row data
            for col_idx in range(1, max_data_cols + 1):
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                cell = sheet[cell_ref]
                cell_value = cell.value

                # Format cell value for display
                display_value = str(cell_value) if cell_value is not None else ""

                # Escape markdown special characters
                display_value = display_value.replace("|", "\\|").replace("\n", " ").replace("\r", " ")

                # Simple format: A1:value
                formatted_cell = f"{cell_ref}:{display_value}"

                row_cells.append(cell_value)
                formatted_row_cells.append(formatted_cell)

            # Convert formatted row to string to estimate tokens
            row_str = " | ".join(formatted_row_cells)
            row_tokens = calculate_token_cost_line(row_str)

            # Check if adding this row would exceed budget
            if tokens_used + row_tokens > token_budget:
                # Try to add at least some rows even if over budget for minimal data
                if rows_shown < 5:  # Ensure we show at least 5 rows if possible
                    preview_data.append(row_cells)
                    formatted_data.append(formatted_row_cells)
                    rows_shown += 1
                    tokens_used += row_tokens
                break

            preview_data.append(row_cells)
            formatted_data.append(formatted_row_cells)
            rows_shown += 1
            tokens_used += row_tokens

        return {
            'data': preview_data,
            'formatted_data': formatted_data,
            'rows_shown': rows_shown,
            'cols_shown': max_data_cols,
            'start_row': start_row,
            'is_truncated': rows_shown < max_data_rows,
            'tokens_used': tokens_used
        }

    def _setup_excel_libraries(self):
        """Setup Excel-related libraries and utilities."""
        try:
            import openpyxl
            from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
            import pandas as pd
            import numpy as np
            import matplotlib
            matplotlib.use('Agg')

            logger.info(f"Loading Excel file: {self.excel_path}")
            start_time = time.time()
            workbook = load_workbook(self.excel_path, data_only=True)
            load_time = time.time() - start_time
            logger.info(f"Excel file loaded in {load_time:.2f}s")
            print(f"📊 [Excel] Loaded in {load_time:.2f}s")

            # Add libraries to code environment
            self.code_globals.update({
                'openpyxl': openpyxl,
                'workbook': workbook,
                'sheet_names': workbook.sheetnames,
                'range_boundaries': range_boundaries,
                'get_column_letter': get_column_letter,
                'column_index_from_string': column_index_from_string,
                'pandas': pd,
                'pd': pd,
                'numpy': np,
                'np': np,
            })

            # Create ExcelToolkit instance and add helper functions
            self.mcp_toolkit = ExcelToolkit(workbook, self.excel_path)
            excel_helpers = self.mcp_toolkit.get_helper_functions_dict()
            self.code_globals.update(excel_helpers)

            logger.info("Excel libraries loaded successfully")
            logger.info(f"Available sheets: {workbook.sheetnames}")
            print("📦 [SheetBrain] Excel libraries loaded successfully")
            print(f"📊 [SheetBrain] Available sheets: {workbook.sheetnames}")

        except ImportError as e:
            logger.error(f"Failed to import required libraries: {e}")
            print(f"❌ [SheetBrain] Failed to import required libraries: {e}")
            raise
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            print(f"❌ [SheetBrain] Failed to load Excel file: {e}")
            raise
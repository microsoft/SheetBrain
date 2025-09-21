# Copyright (c) Microsoft Corporation.
# Licensed under the MIT License.

"""Excel utilities and toolkit for SheetBrain."""

import os
import re
import tempfile
import io
from typing import List, Optional, Dict, Union, Any

import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart
from openpyxl.chart.reference import Reference
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import tiktoken

def calculate_token_cost_line(text: str, model: str = "gpt-4") -> int:
    """
    Calculate the actual token cost of a string using tiktoken.

    Args:
        text: Input text to analyze
        model: Model name for tokenization

    Returns:
        Actual token count
    """
    try:
        # Map model names to tiktoken encodings
        model_encodings = {
            "gpt-4": "cl100k_base",
            "gpt-4-turbo": "cl100k_base",
            "gpt-4o": "o200k_base",
            "gpt-3.5-turbo": "cl100k_base",
            "gpt-5-nano-2025-08-07": "o200k_base", 
            "text-embedding-ada-002": "cl100k_base",
        }

        # Get the appropriate encoding
        encoding_name = model_encodings.get(model, "cl100k_base")  # Default to cl100k_base
        encoding = tiktoken.get_encoding(encoding_name)

        # Encode and count tokens
        tokens = encoding.encode(text)
        return len(tokens)
    
    except Exception:
        # Fallback on any error
        char_count = len(text)
        token_count = max(1, int(char_count / 3.5))
        return token_count


class ExcelToolkit:
    """A comprehensive toolkit for Excel operations with openpyxl."""

    def __init__(self, workbook, excel_path: str):
        """
        Initialize the ExcelToolkit.

        Args:
            workbook: An openpyxl workbook instance
            excel_path: Path to the Excel file
        """
        self.workbook = workbook
        self.excel_path = excel_path
        self._temp_files = []

    def get_sheet(self, sheet_name: Optional[str] = None):
        """Get a worksheet by name or return the active sheet."""
        if sheet_name is None:
            return self.workbook.active
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self.workbook.sheetnames}")

    def inspector(self, range_ref: str, sheet_name: Optional[str] = None) -> List[List]:
        """Read a range of cells and return as list of lists."""
        sheet = self.get_sheet(sheet_name)
        cell_range = sheet[range_ref]

        if hasattr(cell_range, 'value'):
            return [[cell_range.value]]

        result = []
        for row in cell_range:
            row_values = [cell.value for cell in row]
            result.append(row_values)
        return result

    def inspector_attribute(self, range_ref: str, attributes: List[str],
                          sheet_name: Optional[str] = None) -> Dict:
        """Read attributes of a range of cells."""
        print(f"🔎 [read_range_attribute] Reading attributes {attributes} for range {range_ref} in sheet '{sheet_name}'")

        if not attributes:
            return {"error": "No attributes specified"}

        valid_attributes = ["color", "font", "formula"]
        invalid_attrs = [attr for attr in attributes if attr not in valid_attributes]
        if invalid_attrs:
            return {"error": f"Invalid attributes: {invalid_attrs}. Valid options: {valid_attributes}"}

        try:
            sheet = self.get_sheet(sheet_name)
            cell_range = sheet[range_ref]
        except (ValueError, KeyError) as e:
            return {"error": str(e)}

        if hasattr(cell_range, 'coordinate'):
            cells_to_process = [cell_range]
        else:
            cells_to_process = []
            for row in cell_range:
                if hasattr(row, '__iter__'):
                    cells_to_process.extend(row)
                else:
                    cells_to_process.append(row)

        result_attributes = {}

        for attr in attributes:
            result_attributes[attr] = {}

            for cell in cells_to_process:
                cell_coord = cell.coordinate
                attr_value = None

                if attr == "color":
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb != '00000000':
                        attr_value = f"#{cell.fill.fgColor.rgb}"

                elif attr == "font":
                    font_details = []
                    if cell.font:
                        if cell.font.color and cell.font.color.rgb != '00000000':
                            font_details.append(f"color:#{cell.font.color.rgb}")
                        if cell.font.name:
                            font_details.append(f"name:{cell.font.name}")
                        if cell.font.size:
                            font_details.append(f"size:{cell.font.size}")
                        if cell.font.bold:
                            font_details.append("bold:True")
                        if cell.font.italic:
                            font_details.append("italic:True")
                        if cell.font.underline and cell.font.underline != 'none':
                            font_details.append(f"underline:{cell.font.underline}")

                    attr_value = "; ".join(font_details) if font_details else None

                elif attr == "formula":
                    if cell.data_type == 'f' and cell.value:
                        attr_value = str(cell.value)

                if attr_value is not None:
                    result_attributes[attr][cell_coord] = attr_value

        return {
            "range": range_ref,
            "sheet": sheet_name or sheet.title,
            "attributes": result_attributes,
            "total_cells_processed": len(cells_to_process)
        }

    def search(self, value: Any, sheet_name: Optional[str] = None,
              case_sensitive: bool = False, search_type: str = "partial") -> List[Dict]:
        """Find all cells containing a specific value."""
        sheet = self.get_sheet(sheet_name)
        matches = []

        valid_search_types = ["partial", "whole", "strip"]
        if search_type not in valid_search_types:
            raise ValueError(f"Invalid search_type '{search_type}'. Valid options: {valid_search_types}")

        search_value = str(value) if case_sensitive else str(value).lower()

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_str = str(cell.value)

                    if not case_sensitive:
                        cell_str = cell_str.lower()

                    is_match = False

                    if search_type == "partial":
                        is_match = search_value in cell_str
                    elif search_type == "whole":
                        is_match = search_value == cell_str
                    elif search_type == "strip":
                        stripped_cell_str = cell_str.strip()
                        is_match = search_value == stripped_cell_str

                    if is_match:
                        matches.append({
                            'coordinate': cell.coordinate,
                            'value': cell.value,
                            'row': cell.row,
                            'column': cell.column
                        })

        return matches

    def get_sheet_as_dataframe(self, sheet_name: Optional[str] = None,
                              header_row: int = 1, max_rows: Optional[int] = None):
        """Convert a sheet to pandas DataFrame."""
        import pandas as pd
        sheet = self.get_sheet(sheet_name)

        data = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if max_rows and i > max_rows:
                break
            data.append(row)

        if not data:
            return pd.DataFrame()

        if header_row > 0 and len(data) >= header_row:
            headers = data[header_row - 1]
            data_rows = data[header_row:]
            df = pd.DataFrame(data_rows, columns=headers)
        else:
            df = pd.DataFrame(data)

        return df

    def save_plot_to_excel(self, sheet_name: str, cell_position: str = "A1",
                          figsize: tuple = (10, 6), dpi: int = 100) -> str:
        """Save the current matplotlib plot to an Excel sheet."""
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        sheet = self.workbook[sheet_name]

        fig = plt.gcf()
        if fig.get_axes():
            fig.set_size_inches(figsize)
            plt.tight_layout()

            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=dpi, bbox_inches='tight')
            img_buffer.seek(0)

            pil_img = PILImage.open(img_buffer)

            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
                pil_img.save(tmp_file.name, 'PNG')
                tmp_filename = tmp_file.name

            img = Image(tmp_filename)
            sheet.add_image(img, cell_position)

            self._temp_files.append(tmp_filename)

            print(f"✅ Chart saved to sheet '{sheet_name}' at position {cell_position}")
            plt.close(fig)
            return f"Chart saved to {sheet_name}!{cell_position}"
        else:
            print("⚠️ No plot found to save. Create a plot first.")
            return "No plot to save"

    def save_workbook(self) -> str:
        """Save the workbook to file."""
        dir_path = os.path.dirname(self.excel_path)
        base_name = os.path.splitext(os.path.basename(self.excel_path))[0]
        filename = os.path.join(dir_path, f"{base_name}_output.xlsx")

        self.workbook.save(filename)

        # Clean up temporary files
        for temp_file in self._temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                print(f"⚠️ Warning: Could not delete temporary file {temp_file}: {e}")
        self._temp_files = []

        print(f"💾 Workbook saved to: {filename}")
        return filename

    # Excel editing functions
    def insert_rows(self, sheet_name: str, row_index: int, count: int = 1) -> str:
        """Insert empty rows at the specified position."""
        try:
            sheet = self.get_sheet(sheet_name)

            if row_index < 1 or count < 1:
                raise ValueError("Row index and count must be >= 1")

            sheet.insert_rows(row_index, count)

            message = f"✅ Inserted {count} row(s) at row {row_index} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error inserting rows: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def insert_columns(self, sheet_name: str, col_index: Union[int, str], count: int = 1) -> str:
        """Insert empty columns at the specified position."""
        try:
            sheet = self.get_sheet(sheet_name)

            if isinstance(col_index, str):
                col_index = column_index_from_string(col_index)

            if col_index < 1 or count < 1:
                raise ValueError("Column index and count must be >= 1")

            sheet.insert_cols(col_index, count)

            col_letter = get_column_letter(col_index)
            message = f"✅ Inserted {count} column(s) at column {col_letter} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error inserting columns: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def delete_rows(self, sheet_name: str, start_row: int, count: int = 1) -> str:
        """Delete rows starting from the specified position."""
        try:
            sheet = self.get_sheet(sheet_name)

            if start_row < 1 or count < 1:
                raise ValueError("Start row and count must be >= 1")
            if start_row > sheet.max_row:
                raise ValueError(f"Start row {start_row} exceeds sheet max row {sheet.max_row}")

            sheet.delete_rows(start_row, count)

            message = f"✅ Deleted {count} row(s) starting from row {start_row} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error deleting rows: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def delete_columns(self, sheet_name: str, start_col: Union[int, str], count: int = 1) -> str:
        """Delete columns starting from the specified position."""
        try:
            sheet = self.get_sheet(sheet_name)

            if isinstance(start_col, str):
                start_col = column_index_from_string(start_col)

            if start_col < 1 or count < 1:
                raise ValueError("Start column and count must be >= 1")
            if start_col > sheet.max_column:
                raise ValueError(f"Start column {start_col} exceeds sheet max column {sheet.max_column}")

            sheet.delete_cols(start_col, count)

            col_letter = get_column_letter(start_col)
            message = f"✅ Deleted {count} column(s) starting from column {col_letter} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error deleting columns: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def set_cell_value(self, sheet_name: str, cell_ref: str, value: Any) -> str:
        """Set the value of a single cell."""
        try:
            sheet = self.get_sheet(sheet_name)

            if not re.match(r'^[A-Z]+[0-9]+$', cell_ref.upper()):
                raise ValueError(f"Invalid cell reference: {cell_ref}")

            sheet[cell_ref] = value

            message = f"✅ Set cell {cell_ref} to '{value}' in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error setting cell value: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def set_range_values(self, sheet_name: str, start_cell: str,
                        values_2d_array: List[List[Any]]) -> str:
        """Set values for a range of cells using a 2D array."""
        try:
            sheet = self.get_sheet(sheet_name)

            if not re.match(r'^[A-Z]+[0-9]+$', start_cell.upper()):
                raise ValueError(f"Invalid cell reference: {start_cell}")

            if not values_2d_array or not isinstance(values_2d_array, list):
                raise ValueError("values_2d_array must be a non-empty list")

            start_row, start_col = coordinate_to_tuple(start_cell)

            for row_idx, row_values in enumerate(values_2d_array):
                if not isinstance(row_values, list):
                    raise ValueError(f"Row {row_idx} must be a list")

                for col_idx, value in enumerate(row_values):
                    current_row = start_row + row_idx
                    current_col = start_col + col_idx
                    sheet.cell(row=current_row, column=current_col, value=value)

            rows_count = len(values_2d_array)
            cols_count = max(len(row) for row in values_2d_array) if values_2d_array else 0
            end_cell = sheet.cell(row=start_row + rows_count - 1,
                                column=start_col + cols_count - 1).coordinate

            message = f"✅ Set range {start_cell}:{end_cell} ({rows_count}x{cols_count}) in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error setting range values: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def copy_range(self, src_sheet: str, src_range: str, dest_sheet: str, dest_cell: str) -> str:
        """Copy data from one range to another."""
        try:
            src_ws = self.get_sheet(src_sheet)
            dest_ws = self.get_sheet(dest_sheet)

            if ':' not in src_range:
                raise ValueError("Source range must be in format 'A1:B2'")

            source_data = []
            for row in src_ws[src_range]:
                row_data = [cell.value for cell in row]
                source_data.append(row_data)

            if source_data:
                dest_start_row, dest_start_col = coordinate_to_tuple(dest_cell)

                for row_idx, row_values in enumerate(source_data):
                    for col_idx, value in enumerate(row_values):
                        dest_row = dest_start_row + row_idx
                        dest_col = dest_start_col + col_idx
                        dest_ws.cell(row=dest_row, column=dest_col, value=value)

                rows_count = len(source_data)
                cols_count = len(source_data[0]) if source_data else 0
                dest_end_cell = dest_ws.cell(row=dest_start_row + rows_count - 1,
                                           column=dest_start_col + cols_count - 1).coordinate

                message = f"✅ Copied {src_sheet}!{src_range} to {dest_sheet}!{dest_cell}:{dest_end_cell}"
                print(message)
                return message
            else:
                message = "⚠️ No data found in source range"
                print(message)
                return message

        except Exception as e:
            error_msg = f"❌ Error copying range: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def apply_formatting(self, sheet_name: str, range_ref: str, format_dict: Dict[str, Any]) -> str:
        """Apply formatting to a range of cells."""
        try:
            sheet = self.get_sheet(sheet_name)

            if ':' in range_ref:
                cell_range = sheet[range_ref]
                cells = []
                for row in cell_range:
                    if hasattr(row, '__iter__'):
                        cells.extend(row)
                    else:
                        cells.append(row)
            else:
                cells = [sheet[range_ref]]

            for cell in cells:
                if 'fill_color' in format_dict:
                    color = self._parse_color(format_dict['fill_color'])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

                font_kwargs = {}
                if 'font_color' in format_dict:
                    font_kwargs['color'] = self._parse_color(format_dict['font_color'])
                if 'font_size' in format_dict:
                    font_kwargs['size'] = format_dict['font_size']
                if 'font_name' in format_dict:
                    font_kwargs['name'] = format_dict['font_name']
                if 'bold' in format_dict:
                    font_kwargs['bold'] = format_dict['bold']
                if 'italic' in format_dict:
                    font_kwargs['italic'] = format_dict['italic']
                if 'underline' in format_dict:
                    font_kwargs['underline'] = 'single' if format_dict['underline'] else None

                if font_kwargs:
                    cell.font = Font(**font_kwargs)

                if 'border' in format_dict:
                    border_style = format_dict['border']
                    side = Side(style=border_style)
                    cell.border = Border(left=side, right=side, top=side, bottom=side)

                if 'alignment' in format_dict:
                    horizontal = format_dict['alignment']
                    cell.alignment = Alignment(horizontal=horizontal)

            message = f"✅ Applied formatting to {range_ref} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error applying formatting: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def create_chart(self, sheet_name: str, chart_type: str, data_range: str,
                    position: str = "A1", title: str = "",
                    x_axis_title: str = "", y_axis_title: str = "") -> str:
        """Create a chart in the Excel sheet."""
        try:
            sheet = self.get_sheet(sheet_name)

            chart_classes = {
                'bar': BarChart,
                'line': LineChart,
                'pie': PieChart,
                'scatter': ScatterChart,
                'area': AreaChart
            }

            if chart_type.lower() not in chart_classes:
                raise ValueError(f"Unsupported chart type: {chart_type}. Available: {list(chart_classes.keys())}")

            chart_class = chart_classes[chart_type.lower()]
            chart = chart_class()

            if title:
                chart.title = title
            if x_axis_title and hasattr(chart, 'x_axis'):
                chart.x_axis.title = x_axis_title
            if y_axis_title and hasattr(chart, 'y_axis'):
                chart.y_axis.title = y_axis_title

            data = Reference(sheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)

            sheet.add_chart(chart, position)

            message = f"✅ Created {chart_type} chart from {data_range} at {position} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error creating chart: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def add_formula(self, sheet_name: str, cell_ref: str, formula: str) -> str:
        """Add an Excel formula to a cell."""
        try:
            sheet = self.get_sheet(sheet_name)

            if not re.match(r'^[A-Z]+[0-9]+$', cell_ref.upper()):
                raise ValueError(f"Invalid cell reference: {cell_ref}")

            if not formula.startswith('='):
                formula = '=' + formula

            sheet[cell_ref] = formula

            message = f"✅ Added formula '{formula}' to cell {cell_ref} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error adding formula: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def _parse_color(self, color: str) -> str:
        """Parse color from various formats to hex format."""
        color_names = {
            'red': 'FF0000', 'green': '00FF00', 'blue': '0000FF',
            'yellow': 'FFFF00', 'orange': 'FFA500', 'purple': '800080',
            'pink': 'FFC0CB', 'brown': 'A52A2A', 'black': '000000',
            'white': 'FFFFFF', 'gray': '808080', 'grey': '808080'
        }

        if color.startswith('#'):
            return color[1:]
        elif color.lower() in color_names:
            return color_names[color.lower()]
        else:
            return color

    def get_helper_functions_dict(self) -> Dict:
        """Return a dictionary of helper functions for code execution environments."""
        return {
            'get_sheet': self.get_sheet,
            'inspector': self.inspector,
            'inspector_attribute': self.inspector_attribute,
            'search': self.search,
            'get_sheet_as_dataframe': self.get_sheet_as_dataframe,
            'save_plot_to_excel': self.save_plot_to_excel,
            'save_workbook': self.save_workbook,
            #additional editing tools
            'insert_rows': self.insert_rows,
            'insert_columns': self.insert_columns,
            'delete_rows': self.delete_rows,
            'delete_columns': self.delete_columns,
            'set_cell_value': self.set_cell_value,
            'set_range_values': self.set_range_values,
            'copy_range': self.copy_range,
            'apply_formatting': self.apply_formatting,
            'create_chart': self.create_chart,
            'add_formula': self.add_formula
        }